# -*- coding: utf-8 -*-
"""
台股相對輪動模型 - 每日資料更新主程式

流程：
  1. 從 yfinance 抓取最近 N 天的價量資料
  2. 計算配對指標（Ratio、Z-score、Corr、ATR、Vol Regime 等）
  3. 計算週月指標（週MA20斜率、月K脫離整理、週K HL）
  4. 評估進場訊號（10 項條件）
  5. 評估風險訊號（6 項條件）
  6. 寫入儀表板（位置判定 + 建議動作 + 補漲標記）

執行：
  python update_data.py            # 預設抓近 400 個交易日
  python update_data.py --days 800
  python update_data.py --full     # 完全重建
"""
from __future__ import annotations
import sys
import io
import argparse
from pathlib import Path
from datetime import datetime, timedelta

# 強制 stdout 用 UTF-8（避免 Windows console 亂碼）
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\CludeHome\projects\relativeRotation")
XLSX = ROOT / "相對輪動模型.xlsx"
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)

# ------------------------------------------------------------------
# 配對與標的（與 build_workbook.py 同步）
# ------------------------------------------------------------------
PAIRS = [
    {"id": "P01", "theme": "ABF載板",     "a": ("8046", "南電"),   "b": ("3037", "欣興")},
    {"id": "P02", "theme": "AI電源/PSU",  "a": ("2301", "光寶科"), "b": ("2308", "台達電")},
    {"id": "P03", "theme": "AI散熱",      "a": ("3017", "奇鋐"),   "b": ("3324", "雙鴻")},
    {"id": "P04", "theme": "CCL",         "a": ("2383", "台光電"), "b": ("6213", "聯茂")},
    {"id": "P05", "theme": "AI ODM",      "a": ("2382", "廣達"),   "b": ("3231", "緯創")},
    {"id": "P06", "theme": "自行車",      "a": ("9921", "巨大"),   "b": ("9914", "美利達")},
    {"id": "P07", "theme": "精密傳動",    "a": ("2049", "上銀"),   "b": ("4576", "大銀微系統")},
    {"id": "P08", "theme": "DRAM/記憶體", "a": ("2408", "南亞科"), "b": ("2344", "華邦電")},
]
ALL_STOCKS = {
    "8046": "南電", "3037": "欣興",
    "2301": "光寶科", "2308": "台達電",
    "3017": "奇鋐", "3324": "雙鴻",
    "2383": "台光電", "6213": "聯茂",
    "2382": "廣達", "3231": "緯創",
    "9921": "巨大", "9914": "美利達",
    "2049": "上銀", "4576": "大銀微系統",
    "2408": "南亞科", "2344": "華邦電",
}
INDEX_TICKER = "^TWII"

# yfinance 後綴特殊對應（上櫃股用 .TWO）
TICKER_SUFFIX = {
    "3324": ".TWO",  # 雙鴻 (上櫃)
}


# ------------------------------------------------------------------
# 1. 資料抓取（yfinance）
# ------------------------------------------------------------------
def yf_code(code: str) -> str:
    if code.startswith("^"):
        return code
    return f"{code}{TICKER_SUFFIX.get(code, '.TW')}"


def fetch_prices(days: int = 400) -> pd.DataFrame:
    """抓取所有股票 + 大盤的日 K，回傳 long format DataFrame。"""
    end = datetime.now()
    start = end - timedelta(days=int(days * 1.6) + 60)  # 用日曆天往前推，含緩衝

    tickers = [yf_code(c) for c in ALL_STOCKS.keys()] + [INDEX_TICKER]
    print(f"[fetch] 抓取 {len(tickers)} 個標的, 起訖 {start.date()} ~ {end.date()}")

    df_all = yf.download(
        tickers,
        start=start.strftime("%Y-%m-%d"),
        end=(end + timedelta(days=1)).strftime("%Y-%m-%d"),
        auto_adjust=False,
        group_by="ticker",
        progress=False,
        threads=True,
    )

    rows = []
    for t in tickers:
        try:
            sub = df_all[t].dropna(how="all")
        except KeyError:
            print(f"  [warn] 找不到 {t} 的資料")
            continue
        # 同時去除 .TW 與 .TWO 後綴
        code = t.replace(".TWO", "").replace(".TW", "")
        name = "加權指數" if code == "^TWII" else ALL_STOCKS.get(code, code)
        for dt, r in sub.iterrows():
            if pd.isna(r.get("Close")):
                continue
            rows.append({
                "日期": dt.date(),
                "代號": code,
                "名稱": name,
                "開盤": float(r["Open"]) if not pd.isna(r["Open"]) else None,
                "最高": float(r["High"]) if not pd.isna(r["High"]) else None,
                "最低": float(r["Low"]) if not pd.isna(r["Low"]) else None,
                "收盤": float(r["Close"]),
                "成交量": int(r["Volume"]) if not pd.isna(r["Volume"]) else 0,
                "還原收盤": float(r["Adj Close"]) if not pd.isna(r["Adj Close"]) else float(r["Close"]),
            })

    df = pd.DataFrame(rows).sort_values(["代號", "日期"]).reset_index(drop=True)
    print(f"[fetch] 取得 {len(df)} 列資料 ({df['代號'].nunique()} 個標的)")
    # 同時存一份 CSV 作快取
    df.to_csv(DATA_DIR / "prices_daily.csv", index=False, encoding="utf-8-sig")
    return df


# ------------------------------------------------------------------
# 2. 配對指標計算
# ------------------------------------------------------------------
def calc_atr(df: pd.DataFrame, n: int = 14) -> pd.Series:
    high, low, close = df["最高"], df["最低"], df["收盤"]
    prev_close = close.shift(1)
    tr = pd.concat([
        (high - low),
        (high - prev_close).abs(),
        (low - prev_close).abs(),
    ], axis=1).max(axis=1)
    return tr.rolling(n).mean()


def calc_pair_metrics(prices: pd.DataFrame) -> pd.DataFrame:
    """逐配對計算所有指標。"""
    results = []
    by_code = {c: g.set_index("日期").sort_index() for c, g in prices.groupby("代號")}

    for p in PAIRS:
        a_code, _ = p["a"]
        b_code, _ = p["b"]
        if a_code not in by_code or b_code not in by_code:
            print(f"  [warn] 配對 {p['id']} 缺資料，略過")
            continue
        a = by_code[a_code].copy()
        b = by_code[b_code].copy()
        df = pd.DataFrame({
            "A_close": a["收盤"],
            "B_close": b["收盤"],
            "A_high": a["最高"],   "A_low": a["最低"],
            "B_high": b["最高"],   "B_low": b["最低"],
            "A_vol":  a["成交量"], "B_vol":  b["成交量"],
        }).dropna()

        df["Ratio"] = df["A_close"] / df["B_close"]
        df["Ratio_MA5"]  = df["Ratio"].rolling(5).mean()
        df["Ratio_MA20"] = df["Ratio"].rolling(20).mean()
        df["Ratio_MA60"] = df["Ratio"].rolling(60).mean()
        m60  = df["Ratio"].rolling(60).mean()
        s60  = df["Ratio"].rolling(60).std()
        df["Zscore60"] = (df["Ratio"] - m60) / s60
        df["布林上緣2σ"] = m60 + 2 * s60
        df["布林下緣2σ"] = m60 - 2 * s60

        # Correlation (報酬率)
        a_ret = df["A_close"].pct_change()
        b_ret = df["B_close"].pct_change()
        df["Correlation60"] = a_ret.rolling(60).corr(b_ret)

        # ATR
        a_df = pd.DataFrame({"最高": df["A_high"], "最低": df["A_low"], "收盤": df["A_close"]})
        b_df = pd.DataFrame({"最高": df["B_high"], "最低": df["B_low"], "收盤": df["B_close"]})
        df["ATR14_A"] = calc_atr(a_df)
        df["ATR14_B"] = calc_atr(b_df)

        # Volume Regime
        df["A_Vol20"] = df["A_vol"].rolling(20).mean()
        df["A_Vol60"] = df["A_vol"].rolling(60).mean()
        df["VolRegime_A"] = np.where(df["A_Vol20"] > df["A_Vol60"], "資金進場", "資金退潮")
        df["B_Vol20"] = df["B_vol"].rolling(20).mean()
        df["B_Vol60"] = df["B_vol"].rolling(60).mean()
        df["VolRegime_B"] = np.where(df["B_Vol20"] > df["B_Vol60"], "資金進場", "資金退潮")

        # MA20 乖離
        a_ma20 = df["A_close"].rolling(20).mean()
        b_ma20 = df["B_close"].rolling(20).mean()
        df["A_MA20乖離%"] = (df["A_close"] / a_ma20 - 1) * 100
        df["B_MA20乖離%"] = (df["B_close"] / b_ma20 - 1) * 100

        # Relative Strength (累積報酬比較)
        df["RelStrength20"] = (df["A_close"].pct_change(20) - df["B_close"].pct_change(20)) * 100
        df["RelStrength60"] = (df["A_close"].pct_change(60) - df["B_close"].pct_change(60)) * 100

        df = df.reset_index()
        df.insert(1, "配對ID", p["id"])
        results.append(df[[
            "日期", "配對ID",
            "A_close", "B_close",
            "Ratio", "Ratio_MA5", "Ratio_MA20", "Ratio_MA60",
            "Zscore60", "布林上緣2σ", "布林下緣2σ",
            "Correlation60", "ATR14_A", "ATR14_B",
            "A_Vol20", "A_Vol60", "VolRegime_A",
            "B_Vol20", "B_Vol60", "VolRegime_B",
            "A_MA20乖離%", "B_MA20乖離%",
            "RelStrength20", "RelStrength60",
        ]].rename(columns={"A_close": "StockA收盤", "B_close": "StockB收盤"}))

    return pd.concat(results, ignore_index=True) if results else pd.DataFrame()


# ------------------------------------------------------------------
# 3. 週月指標
# ------------------------------------------------------------------
def calc_weekly_monthly(prices: pd.DataFrame) -> pd.DataFrame:
    """計算週MA20斜率、月K脫離整理、週K HL、大盤位置。"""
    by_code = {c: g.set_index("日期").sort_index() for c, g in prices.groupby("代號")}

    twii = by_code.get("^TWII")
    twii_w = None
    twii_wma20 = None
    if twii is not None and len(twii) > 0:
        # dropna 避免春節無交易週導致 rolling 全部 NaN
        twii_w = twii["收盤"].resample("W-FRI").last().dropna()
        twii_wma20 = twii_w.rolling(20).mean()

    rows = []
    for p in PAIRS:
        a_code, _ = p["a"]
        b_code, _ = p["b"]
        if a_code not in by_code or b_code not in by_code:
            continue
        a = by_code[a_code]["收盤"]
        b = by_code[b_code]["收盤"]

        # 週線（dropna 處理假期空週）
        a_w = a.resample("W-FRI").last().dropna()
        b_w = b.resample("W-FRI").last().dropna()
        a_wma20 = a_w.rolling(20).mean()
        b_wma20 = b_w.rolling(20).mean()
        a_slope = a_wma20 - a_wma20.shift(4)
        b_slope = b_wma20 - b_wma20.shift(4)

        # 月線：月K剛突破12個月最高
        a_m = a.resample("ME").last().dropna()
        b_m = b.resample("ME").last().dropna()
        a_break = (a_m > a_m.shift(1).rolling(12).max()).reindex(a.index, method="ffill")
        b_break = (b_m > b_m.shift(1).rolling(12).max()).reindex(b.index, method="ffill")

        # 週K Higher Low：近4週低點 > 前4週低點
        a_w_low = a.resample("W-FRI").min().dropna()
        b_w_low = b.resample("W-FRI").min().dropna()
        a_hl = a_w_low.rolling(4).min() > a_w_low.shift(4).rolling(4).min()
        b_hl = b_w_low.rolling(4).min() > b_w_low.shift(4).rolling(4).min()

        # 重採樣回日線（用 forward-fill）
        a_wma20_d = a_wma20.reindex(a.index, method="ffill")
        b_wma20_d = b_wma20.reindex(b.index, method="ffill")
        a_slope_d = a_slope.reindex(a.index, method="ffill")
        b_slope_d = b_slope.reindex(b.index, method="ffill")
        a_hl_d    = a_hl.reindex(a.index, method="ffill")
        b_hl_d    = b_hl.reindex(b.index, method="ffill")

        if twii_w is not None:
            twii_wma20_d = twii_wma20.reindex(a.index, method="ffill")
            twii_d_close = twii["收盤"].reindex(a.index, method="ffill")
        else:
            twii_wma20_d = pd.Series(np.nan, index=a.index)
            twii_d_close = pd.Series(np.nan, index=a.index)

        df = pd.DataFrame({
            "日期": a.index,
            "配對ID": p["id"],
            "A_週MA20": a_wma20_d.values,
            "A_週MA20斜率": a_slope_d.values,
            "A_週MA20向上": (a_slope_d > 0).values,
            "B_週MA20": b_wma20_d.values,
            "B_週MA20斜率": b_slope_d.values,
            "B_週MA20向上": (b_slope_d > 0).values,
            "A_月K脫離整理": a_break.values,
            "B_月K脫離整理": b_break.values,
            "A_週K_HigherLow": a_hl_d.values,
            "B_週K_HigherLow": b_hl_d.values,
            "加權指數_週MA20": twii_wma20_d.values,
            "加權指數位置": np.where(twii_d_close > twii_wma20_d, "週MA20之上", "週MA20之下"),
        })
        rows.append(df)
    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()


# ------------------------------------------------------------------
# 4. 進場訊號（10 項）
# ------------------------------------------------------------------
def calc_entry_signals(metrics: pd.DataFrame, weekly: pd.DataFrame, prices: pd.DataFrame) -> pd.DataFrame:
    """10 項進場條件逐一評估，輸出每日每配對。"""
    by_code = {c: g.set_index("日期").sort_index() for c, g in prices.groupby("代號")}

    out = []
    for p in PAIRS:
        mid = p["id"]
        m = metrics[metrics["配對ID"] == mid].set_index("日期").sort_index()
        w = weekly[weekly["配對ID"] == mid].set_index("日期").sort_index() if not weekly.empty else pd.DataFrame()
        if m.empty:
            continue

        a_code, _ = p["a"]
        b_code, _ = p["b"]
        a = by_code.get(a_code)
        if a is None:
            continue

        # 條件 ①：Z < -2
        c1 = m["Zscore60"] < -2

        # 條件 ②：低檔翻揚（Z 上升 + Ratio 站回 MA5）
        c2 = (m["Zscore60"] > m["Zscore60"].shift(1)) & (m["Ratio"] > m["Ratio_MA5"])

        # 條件 ③ ④：雙週MA20向上 + 斜率>0（即近4週上升）
        if not w.empty:
            c3 = w["A_週MA20向上"] & w["B_週MA20向上"]
            c4 = (w["A_週MA20斜率"] > 0) & (w["B_週MA20斜率"] > 0)
            c5 = w["A_月K脫離整理"] | w["B_月K脫離整理"]
            c6 = w["A_週K_HigherLow"] & w["B_週K_HigherLow"]
            c10 = w["加權指數位置"] == "週MA20之上"
            c3 = c3.reindex(m.index, method="ffill").fillna(False)
            c4 = c4.reindex(m.index, method="ffill").fillna(False)
            c5 = c5.reindex(m.index, method="ffill").fillna(False)
            c6 = c6.reindex(m.index, method="ffill").fillna(False)
            c10 = c10.reindex(m.index, method="ffill").fillna(False)
        else:
            c3 = c4 = c5 = c6 = c10 = pd.Series(False, index=m.index)

        # 條件 ⑦：日K量縮修正（A收盤 < MA20 且當日量 < 20日均量）
        a_close = a["收盤"]
        a_vol   = a["成交量"]
        a_ma20  = a_close.rolling(20).mean()
        a_vma20 = a_vol.rolling(20).mean()
        c7 = ((a_close < a_ma20) & (a_vol < a_vma20)).reindex(m.index, method="ffill").fillna(False)

        # 條件 ⑧：修正量 < 上漲量
        #   修正日 = 跌日；上漲日 = 漲日，取近 20 日平均量比較
        chg = a_close.pct_change()
        down_vol = a_vol.where(chg < 0).rolling(20, min_periods=5).mean()
        up_vol   = a_vol.where(chg > 0).rolling(20, min_periods=5).mean()
        c8 = (down_vol < up_vol).reindex(m.index, method="ffill").fillna(False)

        # 條件 ⑨：20日均量 > 60日均量
        c9 = (m["A_Vol20"] > m["A_Vol60"]) & (m["B_Vol20"] > m["B_Vol60"])

        flags = pd.DataFrame({
            "①Z<-2": c1,
            "②低檔翻揚": c2,
            "③雙週MA20向上": c3,
            "④週MA20斜率>0": c4,
            "⑤月K脫離整理": c5,
            "⑥週K HigherLow": c6,
            "⑦日K量縮修正": c7,
            "⑧修正量<上漲量": c8,
            "⑨20日量>60日量": c9,
            "⑩大盤站週MA20": c10,
        })
        flags = flags.astype(bool)
        score = flags.sum(axis=1)

        # 進場判定必須以「Ratio 處於低估區（Z<0）」為前提；正 Z 一律不進場
        z = m["Zscore60"]

        def judge_row(i):
            s = int(score.iloc[i])
            zi = float(z.iloc[i]) if not pd.isna(z.iloc[i]) else 0
            if zi >= 0:
                return "× 不適合（Z≥0）"
            if zi < -2 and s >= 8:
                return "✓ 強烈進場"
            if zi < -1.5 and s >= 6:
                return "○ 適合進場"
            if zi < -1 and s >= 4:
                return "△ 觀察"
            return "× 不適合"

        judges = [judge_row(i) for i in range(len(flags))]

        df = flags.copy()
        df.insert(0, "配對ID", mid)
        df["達成項數"] = score
        df["綜合判定"] = judges
        df["備註"] = ""
        df = df.reset_index().rename(columns={"index": "日期"})
        # 統一日期欄名
        if "日期" not in df.columns:
            df.rename(columns={df.columns[0]: "日期"}, inplace=True)
        out.append(df)
    return pd.concat(out, ignore_index=True) if out else pd.DataFrame()


# ------------------------------------------------------------------
# 5. 風險訊號（6 項，拆細到 9 個 flag）
# ------------------------------------------------------------------
def calc_risk_signals(metrics: pd.DataFrame, weekly: pd.DataFrame) -> pd.DataFrame:
    out = []
    for p in PAIRS:
        mid = p["id"]
        m = metrics[metrics["配對ID"] == mid].set_index("日期").sort_index()
        w = weekly[weekly["配對ID"] == mid].set_index("日期").sort_index() if not weekly.empty else pd.DataFrame()
        if m.empty:
            continue

        c1 = m["Correlation60"] < 0.65
        c2 = m["Correlation60"] < 0.5
        c3 = m["A_MA20乖離%"].abs() > 20
        c4 = m["B_MA20乖離%"].abs() > 20

        if not w.empty:
            c5 = ~w["A_週MA20向上"]
            c6 = ~w["B_週MA20向上"]
            c7 = w["加權指數位置"] == "週MA20之下"
            c5 = c5.reindex(m.index, method="ffill").fillna(False)
            c6 = c6.reindex(m.index, method="ffill").fillna(False)
            c7 = c7.reindex(m.index, method="ffill").fillna(False)
        else:
            c5 = c6 = c7 = pd.Series(False, index=m.index)

        # ⑧ Ratio 突破 120日布林上緣 + 爆量（A 或 B 量 > 2x 60 日均量）
        m120 = m["Ratio"].rolling(120).mean()
        s120 = m["Ratio"].rolling(120).std()
        upper120 = m120 + 2 * s120
        a_surge = m["A_Vol20"] / m["A_Vol60"] > 2  # 簡化版
        b_surge = m["B_Vol20"] / m["B_Vol60"] > 2
        c8 = (m["Ratio"] > upper120) & (a_surge | b_surge)

        # ⑨ 時間停損：Z<-2 後 20-30 日仍未回到 -0.5 以內
        z = m["Zscore60"]
        flag = (z < -2)
        # 計算自最近一次 flag=True 起的日數
        idx = np.arange(len(z))
        last_true = pd.Series(np.where(flag, idx, np.nan)).ffill().values
        days_since = idx - last_true
        c9 = (days_since >= 20) & (days_since <= 30) & (z > -0.5).values
        c9 = pd.Series(c9, index=m.index).fillna(False)

        flags = pd.DataFrame({
            "①Corr<0.65": c1,
            "②Corr<0.5": c2,
            "③A_MA20乖離>20%": c3,
            "④B_MA20乖離>20%": c4,
            "⑤A_週MA20轉平/下彎": c5,
            "⑥B_週MA20轉平/下彎": c6,
            "⑦大盤跌破週MA20": c7,
            "⑧Ratio突破布林+爆量": c8,
            "⑨時間停損(20-30日未回歸)": c9,
        }).astype(bool)

        def judge(row):
            if row["②Corr<0.5"] or row["⑦大盤跌破週MA20"] or row["⑧Ratio突破布林+爆量"]:
                return "🛑 停止操作"
            risky = sum([row["①Corr<0.65"], row["③A_MA20乖離>20%"], row["④B_MA20乖離>20%"],
                         row["⑤A_週MA20轉平/下彎"], row["⑥B_週MA20轉平/下彎"], row["⑨時間停損(20-30日未回歸)"]])
            if risky >= 2:
                return "⚠ 降碼"
            return "✓ 繼續"

        flags["綜合判定"] = flags.apply(judge, axis=1)
        flags["備註"] = ""
        flags.insert(0, "配對ID", mid)
        flags = flags.reset_index().rename(columns={"index": "日期"})
        if "日期" not in flags.columns:
            flags.rename(columns={flags.columns[0]: "日期"}, inplace=True)
        out.append(flags)
    return pd.concat(out, ignore_index=True) if out else pd.DataFrame()


# ------------------------------------------------------------------
# 6. 儀表板（取最新一日）
# ------------------------------------------------------------------
def build_dashboard(metrics: pd.DataFrame, entry: pd.DataFrame, risk: pd.DataFrame) -> pd.DataFrame:
    rows = []
    latest_date = metrics["日期"].max() if not metrics.empty else None
    for p in PAIRS:
        mid = p["id"]
        m = metrics[(metrics["配對ID"] == mid) & (metrics["日期"] == latest_date)]
        if m.empty:
            continue
        m = m.iloc[0]
        e = entry[(entry["配對ID"] == mid) & (entry["日期"] == latest_date)]
        r = risk[(risk["配對ID"] == mid) & (risk["日期"] == latest_date)]
        e_judge = e["綜合判定"].iloc[0] if not e.empty else ""
        r_judge = r["綜合判定"].iloc[0] if not r.empty else ""

        # 位置判定
        z = m["Zscore60"]
        a_dev = m["A_MA20乖離%"]
        b_dev = m["B_MA20乖離%"]
        if pd.isna(z):
            phase = "資料不足"
        elif abs(a_dev) > 20 or abs(b_dev) > 20:
            phase = "過熱段"
        elif z < -2:
            phase = "修正段（低估）"
        elif z > 2:
            phase = "過熱段（強勢）"
        elif z < -1:
            phase = "重新轉強"
        elif z < 0:
            phase = "主升初段"
        elif z < 1:
            phase = "主升中段"
        else:
            phase = "主升末段"

        # 建議動作（綜合進場 + 風險判定）
        if "停止" in r_judge or "降碼" in r_judge:
            action = r_judge
        elif "強烈進場" in e_judge or "適合進場" in e_judge:
            action = "建議換股/加碼 StockA"
        elif z > 1.5:
            action = "減碼 StockA / 回補 StockB"
        else:
            action = "觀察"

        # 補漲標記
        about_to_catch_up = (z < -1.5) and ("進場" in e_judge) and ("停止" not in r_judge)

        rows.append({
            "更新日期": latest_date,
            "配對ID": mid,
            "主題": p["theme"],
            "最新Ratio": round(float(m["Ratio"]), 4) if not pd.isna(m["Ratio"]) else None,
            "Z60": round(float(z), 2) if not pd.isna(z) else None,
            "Corr60": round(float(m["Correlation60"]), 3) if not pd.isna(m["Correlation60"]) else None,
            "ATR_A": round(float(m["ATR14_A"]), 2) if not pd.isna(m["ATR14_A"]) else None,
            "ATR_B": round(float(m["ATR14_B"]), 2) if not pd.isna(m["ATR14_B"]) else None,
            "VolRegime_A": m["VolRegime_A"],
            "VolRegime_B": m["VolRegime_B"],
            "A乖離%": round(float(a_dev), 2) if not pd.isna(a_dev) else None,
            "B乖離%": round(float(b_dev), 2) if not pd.isna(b_dev) else None,
            "位置判定": phase,
            "建議動作": action,
            "⭐補漲標記": "★ 即將補漲" if about_to_catch_up else "",
        })
    return pd.DataFrame(rows)


# ------------------------------------------------------------------
# 7. 寫入 Excel（保留標題格式，覆寫資料）
# ------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Microsoft JhengHei", size=11)
BODY_FONT = Font(name="Microsoft JhengHei", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
BAD_FILL  = PatternFill("solid", fgColor="FFC7CE")
HIGHLIGHT = PatternFill("solid", fgColor="FFE699")


def write_sheet(wb, sheet_name: str, df: pd.DataFrame, color_rules: dict = None):
    """覆寫整個分頁的內容（保留第1列標題）。"""
    ws = wb[sheet_name]
    # 清除舊資料（保留 row 1）
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    if df is None or df.empty:
        return

    headers = [c.value for c in ws[1]]
    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        for c_idx, h in enumerate(headers, 1):
            val = row.get(h, None)
            if isinstance(val, (np.bool_, bool)):
                val = bool(val)
            if pd.isna(val) if val is not None and not isinstance(val, str) else False:
                val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            # 顏色規則
            if color_rules and h in color_rules:
                fill = color_rules[h](val, row)
                if fill is not None:
                    cell.fill = fill


def dashboard_color(val, row):
    phase = row.get("位置判定", "")
    if "過熱" in phase: return BAD_FILL
    if "低估" in phase or "轉強" in phase: return GOOD_FILL
    if "修正" in phase: return WARN_FILL
    return None


def entry_color(val, row):
    j = row.get("綜合判定", "")
    if "強烈進場" in j: return GOOD_FILL
    if "適合進場" in j: return HIGHLIGHT
    if "觀察" in j: return WARN_FILL
    return None


def risk_color(val, row):
    j = row.get("綜合判定", "")
    if "停止" in j: return BAD_FILL
    if "降碼" in j: return WARN_FILL
    if "繼續" in j: return GOOD_FILL
    return None


def update_workbook(prices, metrics, weekly, entry, risk, dashboard):
    if not XLSX.exists():
        print(f"[error] 找不到 {XLSX}，請先執行 build_workbook.py")
        return
    wb = load_workbook(XLSX)
    write_sheet(wb, "每日價格", prices)
    write_sheet(wb, "配對指標", metrics)
    write_sheet(wb, "週月指標", weekly)
    write_sheet(wb, "進場訊號", entry,
                color_rules={"綜合判定": entry_color})
    write_sheet(wb, "風險訊號", risk,
                color_rules={"綜合判定": risk_color})
    write_sheet(wb, "儀表板", dashboard,
                color_rules={
                    "位置判定": dashboard_color,
                    "⭐補漲標記": lambda v, r: HIGHLIGHT if v else None,
                    "建議動作": lambda v, r: (BAD_FILL if "停止" in str(v) else
                                              WARN_FILL if "減碼" in str(v) or "降碼" in str(v) else
                                              GOOD_FILL if "加碼" in str(v) or "換股" in str(v) else None),
                })
    wb.save(XLSX)
    print(f"[ok] 寫入完成: {XLSX}")


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--days", type=int, default=400, help="抓取近 N 個交易日 (預設 400)")
    ap.add_argument("--full", action="store_true", help="完全重抓（同 --days 800）")
    args = ap.parse_args()

    days = 800 if args.full else args.days
    print(f"=== 相對輪動模型更新開始 ({datetime.now():%Y-%m-%d %H:%M:%S}) ===")

    prices = fetch_prices(days=days)
    if prices.empty:
        print("[error] 無資料")
        return

    # 過濾掉太舊的列，避免 Excel 太大（保留近 days 個交易日）
    cutoff = prices["日期"].max() - timedelta(days=int(days * 1.6))
    prices = prices[prices["日期"] >= cutoff].reset_index(drop=True)

    # 轉成 datetime index 給後續計算
    prices["日期"] = pd.to_datetime(prices["日期"])

    metrics = calc_pair_metrics(prices)
    weekly  = calc_weekly_monthly(prices)
    entry   = calc_entry_signals(metrics, weekly, prices)
    risk    = calc_risk_signals(metrics, weekly)
    dash    = build_dashboard(metrics, entry, risk)

    # 還原成 date（Excel 顯示）
    for df in (prices, metrics, weekly, entry, risk):
        if "日期" in df.columns:
            df["日期"] = pd.to_datetime(df["日期"]).dt.date
    if "更新日期" in dash.columns:
        dash["更新日期"] = pd.to_datetime(dash["更新日期"]).dt.date

    # 只保留每張表近 N 列（避免太大）
    keep_days = days
    metrics_recent = metrics.groupby("配對ID").tail(keep_days)
    weekly_recent  = weekly.groupby("配對ID").tail(keep_days)
    entry_recent   = entry.groupby("配對ID").tail(keep_days)
    risk_recent    = risk.groupby("配對ID").tail(keep_days)

    update_workbook(prices, metrics_recent, weekly_recent, entry_recent, risk_recent, dash)

    # 印出今日儀表板摘要
    print("\n=== 最新儀表板 ===")
    if not dash.empty:
        cols_show = ["配對ID", "主題", "最新Ratio", "Z60", "Corr60",
                     "A乖離%", "B乖離%", "位置判定", "建議動作", "⭐補漲標記"]
        print(dash[cols_show].to_string(index=False))
    print(f"\n=== 完成 ({datetime.now():%H:%M:%S}) ===")


if __name__ == "__main__":
    main()
