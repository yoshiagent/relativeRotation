# -*- coding: utf-8 -*-
"""
建立『相對輪動模型.xlsx』主工作簿
共 9 個分頁，預填配對池與股票主檔。
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\CludeHome\projects\relativeRotation")
XLSX = ROOT / "相對輪動模型.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Microsoft JhengHei", size=11)
BODY_FONT = Font(name="Microsoft JhengHei", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def add_headers(ws, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
        if widths and i - 1 < len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.freeze_panes = "A2"


def fill_row(ws, row, values):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = BORDER


def main():
    wb = Workbook()
    wb.remove(wb.active)

    # 1. 配對池
    ws = wb.create_sheet("配對池")
    add_headers(ws, [
        "配對ID", "主題",
        "StockA代號", "StockA名稱",
        "StockB代號", "StockB名稱",
        "啟用狀態", "加入日期", "備註",
    ], widths=[10, 16, 12, 14, 12, 14, 10, 14, 30])
    pairs = [
        ("P01", "ABF載板",     "8046", "南電",   "3037", "欣興",   "✓", "2026-05-20", "南電vs欣興"),
        ("P02", "AI電源/PSU",  "2301", "光寶科", "2308", "台達電", "✓", "2026-05-20", "光寶vs台達電 - BBU/HVDC"),
        ("P03", "AI散熱",      "3017", "奇鋐",   "3324", "雙鴻",   "✓", "2026-05-20", "奇鋐vs雙鴻"),
        ("P04", "CCL",         "2383", "台光電", "6213", "聯茂",   "✓", "2026-05-20", "台光電vs聯茂"),
        ("P05", "AI ODM",      "2382", "廣達",   "3231", "緯創",   "✓", "2026-05-20", "廣達vs緯創"),
    ]
    for i, r in enumerate(pairs, 2):
        fill_row(ws, i, r)

    # 2. 股票主檔
    ws = wb.create_sheet("股票主檔")
    add_headers(ws, [
        "代號", "yfinance代碼", "名稱", "產業", "所屬主題", "上市別", "加入日",
    ], widths=[10, 14, 14, 14, 16, 10, 14])
    stocks = [
        ("8046", "8046.TW", "南電",   "電子零組件", "ABF載板",     "上市", "2026-05-20"),
        ("3037", "3037.TW", "欣興",   "電子零組件", "ABF載板",     "上市", "2026-05-20"),
        ("2301", "2301.TW", "光寶科", "電子零組件", "AI電源/PSU",  "上市", "2026-05-20"),
        ("2308", "2308.TW", "台達電", "電子零組件", "AI電源/PSU",  "上市", "2026-05-20"),
        ("3017", "3017.TW", "奇鋐",   "電子零組件", "AI散熱",      "上市", "2026-05-20"),
        ("3324", "3324.TWO", "雙鴻",  "電子零組件", "AI散熱",      "上櫃", "2026-05-20"),
        ("2383", "2383.TW", "台光電", "電子零組件", "CCL",         "上市", "2026-05-20"),
        ("6213", "6213.TW", "聯茂",   "電子零組件", "CCL",         "上市", "2026-05-20"),
        ("2382", "2382.TW", "廣達",   "電腦及週邊", "AI ODM",      "上市", "2026-05-20"),
        ("3231", "3231.TW", "緯創",   "電腦及週邊", "AI ODM",      "上市", "2026-05-20"),
        ("^TWII", "^TWII", "加權指數", "指數",       "大盤",        "指數", "2026-05-20"),
    ]
    for i, r in enumerate(stocks, 2):
        fill_row(ws, i, r)

    # 3. 每日價格
    ws = wb.create_sheet("每日價格")
    add_headers(ws, [
        "日期", "代號", "名稱", "開盤", "最高", "最低", "收盤", "成交量", "還原收盤",
    ], widths=[12, 10, 12, 10, 10, 10, 10, 14, 12])

    # 4. 配對指標
    ws = wb.create_sheet("配對指標")
    add_headers(ws, [
        "日期", "配對ID",
        "StockA收盤", "StockB收盤",
        "Ratio", "Ratio_MA5", "Ratio_MA20", "Ratio_MA60",
        "Zscore60", "布林上緣2σ", "布林下緣2σ",
        "Correlation60", "ATR14_A", "ATR14_B",
        "A_Vol20", "A_Vol60", "VolRegime_A",
        "B_Vol20", "B_Vol60", "VolRegime_B",
        "A_MA20乖離%", "B_MA20乖離%",
        "RelStrength20", "RelStrength60",
    ], widths=[12, 8] + [12] * 22)

    # 5. 週月指標
    ws = wb.create_sheet("週月指標")
    add_headers(ws, [
        "日期", "配對ID",
        "A_週MA20", "A_週MA20斜率", "A_週MA20向上",
        "B_週MA20", "B_週MA20斜率", "B_週MA20向上",
        "A_月K脫離整理", "B_月K脫離整理",
        "A_週K_HigherLow", "B_週K_HigherLow",
        "加權指數_週MA20", "加權指數位置",
    ], widths=[12, 8] + [14] * 12)

    # 6. 進場訊號
    ws = wb.create_sheet("進場訊號")
    add_headers(ws, [
        "日期", "配對ID",
        "①Z<-2", "②低檔翻揚", "③雙週MA20向上", "④週MA20斜率>0",
        "⑤月K脫離整理", "⑥週K HigherLow", "⑦日K量縮修正",
        "⑧修正量<上漲量", "⑨20日量>60日量", "⑩大盤站週MA20",
        "達成項數", "綜合判定", "備註",
    ], widths=[12, 8] + [12] * 10 + [10, 14, 24])

    # 7. 風險訊號
    ws = wb.create_sheet("風險訊號")
    add_headers(ws, [
        "日期", "配對ID",
        "①Corr<0.65", "②Corr<0.5",
        "③A_MA20乖離>20%", "④B_MA20乖離>20%",
        "⑤A_週MA20轉平/下彎", "⑥B_週MA20轉平/下彎",
        "⑦大盤跌破週MA20",
        "⑧Ratio突破布林+爆量", "⑨時間停損(20-30日未回歸)",
        "綜合判定", "備註",
    ], widths=[12, 8] + [16] * 9 + [14, 24])

    # 8. 持倉與交易記錄
    ws = wb.create_sheet("持倉與交易記錄")
    add_headers(ws, [
        "日期", "配對ID", "動作", "標的代號", "標的名稱",
        "張數", "成交價", "成交金額", "累計部位", "理由",
    ], widths=[12, 8, 12, 10, 12, 10, 10, 14, 14, 30])

    # 9. 儀表板
    ws = wb.create_sheet("儀表板")
    add_headers(ws, [
        "更新日期", "配對ID", "主題",
        "最新Ratio", "Z60", "Corr60",
        "ATR_A", "ATR_B",
        "VolRegime_A", "VolRegime_B",
        "A乖離%", "B乖離%",
        "位置判定", "建議動作", "⭐補漲標記",
    ], widths=[12, 8, 14, 10, 8, 8, 8, 8, 12, 12, 10, 10, 14, 14, 14])

    wb.save(XLSX)
    print(f"已建立: {XLSX}")
    print(f"分頁數: {len(wb.sheetnames)} -> {wb.sheetnames}")


if __name__ == "__main__":
    main()
